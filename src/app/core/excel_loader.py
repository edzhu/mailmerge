"""Excel data loading utilities for the mail-merge emailer."""

from __future__ import annotations

from datetime import date, datetime
from numbers import Number
from pathlib import Path
from typing import Any, Dict, Mapping, Optional, Sequence, TYPE_CHECKING

from app.core.canonicalize import canonicalize
from app.core.errors import ExcelValidationError, OptionalDependencyError
from app.core.models import (
    CellValue,
    ColumnInfo,
    FieldInfo,
    RowData,
    SheetInfo,
    TemplateInfo,
)

if TYPE_CHECKING:
    from openpyxl.cell.read_only import ReadOnlyCell
    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

_HEADER_SCAN_LIMIT = 50
_MIN_HEADER_NON_EMPTY = 2


def load_matching_sheet(
    excel_path: Path, template: TemplateInfo
) -> tuple[SheetInfo, list[RowData]]:
    """Load the first worksheet whose headers satisfy the template."""

    workbook = _load_workbook(Path(excel_path))
    try:
        required_keys = _required_template_keys(template)
        template_fields = _template_fields_by_key(template)
        for sheet in workbook.worksheets:
            header_match = _find_header_row(sheet)
            if header_match is None:
                continue
            header_row_index, header_cells = header_match
            columns = _build_columns(header_cells, template_fields)
            header_keys = {column.key for column in columns}
            if required_keys.issubset(header_keys):
                sheet_info = SheetInfo(
                    name=sheet.title,
                    columns=columns,
                    header_row_index=header_row_index,
                )
                rows = _extract_rows(sheet, sheet_info)
                if rows is None:
                    rows = []
                if len(rows) == 0:
                    raise ExcelValidationError("No data rows found.")
                return sheet_info, rows
        raise ExcelValidationError("No sheet contains all required template fields")
    finally:
        workbook.close()


def normalize_cell_value(value: object) -> CellValue:
    """Normalize a raw Excel cell value."""

    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value
    if isinstance(value, str):
        return value
    if isinstance(value, Number):
        return value
    return str(value)


def _load_openpyxl() -> Any:
    """Load openpyxl and translate missing dependency errors."""

    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise OptionalDependencyError(
            "openpyxl is required to load Excel files."
        ) from exc
    return openpyxl


def _load_workbook(excel_path: Path) -> "Workbook":
    """Load the Excel workbook in read-only mode."""

    openpyxl = _load_openpyxl()
    try:
        return openpyxl.load_workbook(
            excel_path,
            data_only=True,
            read_only=True,
        )
    except FileNotFoundError as exc:
        raise ExcelValidationError(f"Excel file not found: {excel_path}") from exc
    except openpyxl.utils.exceptions.InvalidFileException as exc:
        raise ExcelValidationError("Excel file is not a valid .xlsx file.") from exc
    except OSError as exc:
        raise ExcelValidationError(f"Unable to read Excel file: {excel_path}") from exc


def _required_template_keys(template: TemplateInfo) -> set[str]:
    """Return the canonical required field keys for the template."""

    required_keys: set[str] = set()
    for field in template.fields:
        key = canonicalize(field.key)
        if key and field.required:
            required_keys.add(key)
    return required_keys


def _template_fields_by_key(template: TemplateInfo) -> Dict[str, FieldInfo]:
    """Return template fields mapped by canonical key."""

    fields: Dict[str, FieldInfo] = {}
    for field in template.fields:
        key = canonicalize(field.key)
        if key and key not in fields:
            fields[key] = field
    return fields


def _find_header_row(
    sheet: "Worksheet",
) -> Optional[tuple[int, Sequence["ReadOnlyCell"]]]:
    """Find the first candidate header row in the sheet."""

    for row in sheet.iter_rows(min_row=1, max_row=_HEADER_SCAN_LIMIT):
        if _row_has_minimum_values(row, _MIN_HEADER_NON_EMPTY):
            row_index = row[0].row if row else 1
            return row_index, row
    return None


def _row_has_minimum_values(row: Sequence[Any], minimum: int) -> bool:
    count = 0
    for cell in row:
        if _cell_has_value(cell.value):
            count += 1
            if count >= minimum:
                return True
    return False


def _cell_has_value(value: object) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _build_columns(
    header_cells: Sequence[Any],
    template_fields: Mapping[str, FieldInfo],
) -> list[ColumnInfo]:
    columns: list[ColumnInfo] = []
    seen_keys: set[str] = set()
    for cell in header_cells:
        header_text = _header_text(cell.value)
        if header_text is None:
            continue
        key = canonicalize(header_text)
        if not key or key in seen_keys:
            continue
        field = template_fields.get(key)
        columns.append(
            ColumnInfo(
                index=int(cell.column),
                header=header_text,
                key=key,
                required=bool(field.required) if field else False,
            )
        )
        seen_keys.add(key)
    return columns


def _header_text(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return text


def _extract_rows(sheet: "Worksheet", sheet_info: SheetInfo) -> list[RowData]:
    """Extract row data for the given sheet."""

    if not sheet_info.columns:
        return []
    max_column = max(column.index for column in sheet_info.columns)
    rows: list[RowData] = []
    start_row = sheet_info.header_row_index + 1
    for row_index, row in enumerate(
        sheet.iter_rows(min_row=start_row, max_col=max_column),
        start=start_row,
    ):
        values_by_key: Dict[str, CellValue] = {}
        has_data = False
        for column in sheet_info.columns:
            cell = row[column.index - 1] if column.index - 1 < len(row) else None
            value = normalize_cell_value(cell.value if cell is not None else None)
            if _cell_has_value(value):
                has_data = True
            values_by_key[column.key] = value
        if not has_data:
            continue
        rows.append(RowData(row_index=row_index, values_by_key=values_by_key))
    return rows
